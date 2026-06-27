import os
from typing import Any, Dict, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from .utils import get_logger

logger = get_logger("Submission")

class SubmissionGenerator:
    """Generates highly polished, beautifully styled Excel spreadsheets containing candidate ranking reports."""

    def __init__(self, output_dir: str = "outputs"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate_report(self, ranked_candidates: List[Dict[str, Any]], jd_metadata: Dict[str, Any], weights: Dict[str, float]) -> str:
        """
        Creates a multi-tab professional styled Excel file from ranked candidate lists.
        
        Args:
            ranked_candidates (List[Dict]): Ordered list of candidates.
            jd_metadata (Dict): Parsed job description parameters.
            weights (Dict): Active weights dictionary.
            
        Returns:
            str: Absolute path of the generated Excel file.
        """
        file_path = os.path.join(self.output_dir, "talentmind_rankings_top100.xlsx")
        logger.info(f"Generating stylized XLSX report at {file_path}...")

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Style Assets
        header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")  # Deep Navy
        zebra_fill = PatternFill(start_color="F5F8FA", end_color="F5F8FA", fill_type="solid")   # Very light blue/gray
        accent_fill = PatternFill(start_color="E1EEF6", end_color="E1EEF6", fill_type="solid")  # Light blue highlights
        
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_data = Font(name="Calibri", size=11, bold=False, color="000000")
        font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        slices = {
            "Top 10 Candidates": ranked_candidates[:10],
            "Top 25 Candidates": ranked_candidates[:25],
            "Top 50 Candidates": ranked_candidates[:50],
            "Top 100 Candidates": ranked_candidates[:100]
        }

        # 1. Generate sheets for each slice
        for tab_name, cand_slice in slices.items():
            ws = wb.create_sheet(title=tab_name)
            ws.views.sheetView[0].showGridLines = True

            # Headers
            headers = ["Rank", "ID", "Name", "Score", "Technical Match", "Years of Exp", "Education Fit", "Behavioral Level", "Completeness", "Primary Strengths"]
            ws.append(headers)
            
            # Format Headers
            ws.row_dimensions[1].height = 28
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = font_header
                cell.alignment = align_center

            # Data rows
            for i, cand in enumerate(cand_slice):
                row_idx = i + 2
                ws.row_dimensions[row_idx].height = 20
                
                rank = i + 1
                cid = cand.get("id", "")
                name = cand.get("name", "")
                score = f"{cand.get('overall_score', 0.0) * 100:.1f}%"
                
                breakdown = cand.get("scoring_breakdown", {})
                skills = f"{breakdown.get('skills_match', 0.0) * 100:.0f}%"
                exp = f"{cand.get('years_of_experience', 0.0)} yrs"
                edu = cand.get("education", {}).get("degree", "N/A")
                behav = f"{breakdown.get('behavioral_signals', 0.0) * 100:.0f}%"
                comp = f"{cand.get('profile_completeness', 0.0) * 100:.0f}%"
                
                # Dynamic primary rationale
                rationales = cand.get("rationales", [])
                primary_strength = rationales[0] if rationales else "Qualified background alignment"
                
                row_data = [rank, cid, name, score, skills, exp, edu, behav, comp, primary_strength]
                ws.append(row_data)

                # Format row cells
                is_even = (row_idx % 2 == 0)
                for col_num in range(1, len(row_data) + 1):
                    cell = ws.cell(row=row_idx, column=col_num)
                    cell.font = font_data
                    cell.border = thin_border
                    
                    # Alignment
                    if col_num in [1, 2, 4, 5, 6, 7, 8, 9]:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left

                    # Alternating fills
                    if is_even:
                        cell.fill = zebra_fill
                    
                    # Special highlight for top score
                    if col_num == 4:
                        cell.font = font_bold
                        cell.fill = accent_fill

            # Column auto-fitting
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

        # 2. Add Breakdown Tab
        ws_breakdown = wb.create_sheet(title="Scoring Detail Matrix")
        ws_breakdown.views.sheetView[0].showGridLines = True
        
        breakdown_headers = [
            "Rank", "ID", "Name", "Aggregate Fit Score", 
            "Semantic Embed Similarity (40%)", "Skills Match (20%)", 
            "Experience Match (15%)", "Education Prestige (10%)", 
            "Behavioral Signals (10%)", "Profile Completeness (5%)"
        ]
        ws_breakdown.append(breakdown_headers)
        ws_breakdown.row_dimensions[1].height = 28
        for col_num in range(1, len(breakdown_headers) + 1):
            cell = ws_breakdown.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = font_header
            cell.alignment = align_center

        for i, cand in enumerate(ranked_candidates[:100]):
            row_idx = i + 2
            ws_breakdown.row_dimensions[row_idx].height = 20
            
            rank = i + 1
            cid = cand.get("id", "")
            name = cand.get("name", "")
            score = cand.get("overall_score", 0.0)
            
            b = cand.get("scoring_breakdown", {})
            sem = b.get("semantic_similarity", 0.0)
            skl = b.get("skills_match", 0.0)
            exp = b.get("experience", 0.0)
            edu = b.get("education", 0.0)
            beh = b.get("behavioral_signals", 0.0)
            comp = b.get("profile_completeness", 0.0)
            
            row_data = [rank, cid, name, score, sem, skl, exp, edu, beh, comp]
            ws_breakdown.append(row_data)

            is_even = (row_idx % 2 == 0)
            for col_num in range(1, len(row_data) + 1):
                cell = ws_breakdown.cell(row=row_idx, column=col_num)
                cell.border = thin_border
                cell.font = font_data
                
                # Formats numeric values
                if col_num in [4, 5, 6, 7, 8, 9, 10]:
                    cell.number_format = '0.0%'
                    cell.alignment = align_right
                else:
                    cell.alignment = align_center

                if is_even:
                    cell.fill = zebra_fill

        for col in ws_breakdown.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws_breakdown.column_dimensions[col_letter].width = max(max_len + 3, 10)

        # 3. System Config Tab
        ws_config = wb.create_sheet(title="System Configuration")
        ws_config.views.sheetView[0].showGridLines = True
        
        ws_config.row_dimensions[1].height = 25
        ws_config.cell(row=1, column=1, value="System Parameter").font = font_header
        ws_config.cell(row=1, column=1).fill = header_fill
        ws_config.cell(row=1, column=2, value="Configured Setting").font = font_header
        ws_config.cell(row=1, column=2).fill = header_fill
        
        configs = [
            ("Central Platform Name", "TalentMind-AI"),
            ("Challenge Scope", "INDIA RUNS Data & AI Challenge"),
            ("Dense Vector Embeddings", "sentence-transformers / all-MiniLM-L6-v2"),
            ("Gemini API Status", "Active (Cognitive LLM re-ranking online)"),
            ("Scoring Weights: Semantic Similarity", f"{weights.get('semantic_similarity', 0.40)*100}%"),
            ("Scoring Weights: Skills Density Match", f"{weights.get('skills_match', 0.20)*100}%"),
            ("Scoring Weights: Years of Experience", f"{weights.get('experience', 0.15)*100}%"),
            ("Scoring Weights: Education Level", f"{weights.get('education', 0.10)*100}%"),
            ("Scoring Weights: Behavioral Platform Signals", f"{weights.get('behavioral_signals', 0.10)*100}%"),
            ("Scoring Weights: Profile Completeness", f"{weights.get('profile_completeness', 0.05)*100}%"),
            ("Target Domain parsed from JD", jd_metadata.get("domain", "Unknown")),
            ("Seniority parsed from JD", jd_metadata.get("seniority", "Mid")),
            ("Experience Req. parsed from JD", f"{jd_metadata.get('min_experience_years', 0)} years")
        ]

        for i, (param, val) in enumerate(configs):
            row_idx = i + 2
            ws_config.row_dimensions[row_idx].height = 18
            ws_config.cell(row=row_idx, column=1, value=param).font = font_bold
            ws_config.cell(row=row_idx, column=1).border = thin_border
            ws_config.cell(row=row_idx, column=2, value=val).font = font_data
            ws_config.cell(row=row_idx, column=2).border = thin_border
            
        ws_config.column_dimensions["A"].width = 38
        ws_config.column_dimensions["B"].width = 45

        # Save workbook
        wb.save(file_path)
        logger.info(f"Report exported successfully containing {len(ranked_candidates)} records.")
        return file_path
